# apps/staff/views.py
from django.shortcuts import get_object_or_404
from django.utils.dateparse import parse_datetime
from django.http import StreamingHttpResponse, HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views import View

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser

import io
import csv
import json

from .models import Staff
from apps.catalog.models import MenuItem
from .auth import StaffJWTAuthentication, issue_access_token, set_auth_cookie, clear_auth_cookie
from .permissions import IsOwnerOrManager
from .serializers import (
    StaffLoginSerializer,
    CouponSerializer, MembershipSerializer,
    StaffMeSerializer,
    StaffOrderDetailSerializer,
    InventoryItemUpdateSerializer,
    InventoryItemPartialUpdateSerializer,
)
from apps.promotion.models import Coupon, Membership
from apps.orders.models import Order
from .eventbus import iter_order_notifications

# ===== drf-spectacular =====
from drf_spectacular.utils import (
    extend_schema, OpenApiParameter, OpenApiExample, OpenApiResponse, inline_serializer
)
from rest_framework import serializers


# ---------- Auth ----------
@extend_schema(
    tags=["Staff/Auth"],
    summary="스태프 로그인",
    description=(
        "성공 시 JSON과 함께 **HTTPOnly 쿠키**가 설정됩니다. "
        "프런트엔드는 **Authorization 헤더 없이 쿠키 기반 인증**을 사용합니다."
    ),
    request=StaffLoginSerializer,
    responses={
        200: inline_serializer(
            name="StaffLoginResp",
            fields={
                "status": serializers.BooleanField(),
                "staff_id": serializers.IntegerField(),
            }
        ),
        400: OpenApiResponse(description="아이디 또는 비밀번호 불일치"),
    },
    examples=[
        OpenApiExample(
            name="요청 예시",
            value={"username": "manager1", "password": "P@ssw0rd-1"},
            request_only=True
        ),
        OpenApiExample(
            name="응답 예시(성공)",
            value={"status": True, "staff_id": 3},
            response_only=True
        ),
        OpenApiExample(
            name="응답 예시(실패)",
            value={"detail": "아이디 또는 비밀번호가 올바르지 않습니다."},
            response_only=True
        ),
    ]
)
class StaffLoginView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        ser = StaffLoginSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        username = (ser.validated_data["username"] or "").strip()
        password = ser.validated_data["password"]

        staff = Staff.objects.filter(username__iexact=username).first()
        if not staff or not staff.check_password(password):
            return Response({"detail": "아이디 또는 비밀번호가 올바르지 않습니다."},
                            status=status.HTTP_400_BAD_REQUEST)

        token = issue_access_token(staff)
        resp = Response({"status": True, "staff_id": staff.pk})
        set_auth_cookie(resp, token)
        return resp


@extend_schema(
    tags=["Staff/Auth"],
    summary="스태프 로그아웃",
    description="인증 쿠키를 제거합니다.",
    responses=inline_serializer(
        name="StaffLogoutResp",
        fields={"status": serializers.BooleanField()}
    ),
    examples=[OpenApiExample(name="응답 예시", value={"status": True}, response_only=True)]
)
class StaffLogoutView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        resp = Response({"status": True})
        clear_auth_cookie(resp)
        return resp


@extend_schema(
    tags=["Staff/Auth"],
    summary="내 스태프 정보",
    responses=StaffMeSerializer,
    examples=[OpenApiExample(
        name="응답 예시",
        value={"id": 3, "username": "manager1", "role": "MANAGER", "is_active": True, "created_at": "2025-10-01T10:00:00+09:00"},
        response_only=True
    )]
)
class StaffMeView(APIView):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(StaffMeSerializer(request.user).data)


# ---------- Coupons ----------
@extend_schema(
    methods=["GET"],
    tags=["Staff/Coupons"],
    summary="쿠폰 목록",
    responses=CouponSerializer,  # list 응답은 스펙타큘러가 배열로 표시
    examples=[OpenApiExample(
        name="응답 예시",
        value=[{"code": "WELCOME10", "name": "신규 10%", "label": "WELCOME10", "active": True, "kind": "percent",
                "value": 10, "valid_from": "2025-01-01T00:00:00+09:00", "valid_until": None,
                "min_subtotal_cents": 0, "max_discount_cents": 5000,
                "stackable_with_membership": False, "stackable_with_coupons": False, "channel": "GUI",
                "max_redemptions_global": None, "max_redemptions_per_user": 1,
                "notes": None, "created_at": "2025-01-01T00:00:00+09:00", "updated_at": "2025-01-01T00:00:00+09:00"}],
        response_only=True
    )]
)
@extend_schema(
    methods=["POST"],
    tags=["Staff/Coupons"],
    summary="쿠폰 생성",
    request=CouponSerializer,
    responses={201: CouponSerializer, 403: OpenApiResponse(description="권한 없음 (OWNER/MANAGER만)")},
)
class CouponsView(APIView):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Coupon.objects.all().order_by("-valid_from", "code")
        return Response(CouponSerializer(qs, many=True).data)

    def post(self, request):
        if not IsOwnerOrManager().has_permission(request, self):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        ser = CouponSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        obj = ser.save()
        return Response(CouponSerializer(obj).data, status=status.HTTP_201_CREATED)


@extend_schema(
    tags=["Staff/Coupons"],
    summary="쿠폰 단건 조회/수정/삭제",
    parameters=[OpenApiParameter("code", str, OpenApiParameter.PATH, description="쿠폰 코드(대소문자 무시)")],
    responses={
        200: CouponSerializer,
        204: OpenApiResponse(description="삭제 성공"),
        400: OpenApiResponse(description="잘못된 요청"),
        403: OpenApiResponse(description="권한 없음 (수정/삭제)"),
        404: OpenApiResponse(description="존재하지 않음"),
    },
)
class CouponDetailView(APIView):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_object(self, code: str) -> Coupon:
        return get_object_or_404(Coupon, code=code.upper())

    def get(self, request, code: str):
        return Response(CouponSerializer(self.get_object(code)).data)

    def patch(self, request, code: str):
        if not IsOwnerOrManager().has_permission(request, self):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        obj = self.get_object(code)
        ser = CouponSerializer(instance=obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(CouponSerializer(obj).data)

    def delete(self, request, code: str):
        if not IsOwnerOrManager().has_permission(request, self):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        obj = self.get_object(code)
        if hasattr(obj, "active"):
            obj.active = False
            obj.save(update_fields=["active"])
            return Response(status=status.HTTP_204_NO_CONTENT)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ---------- Memberships ----------
@extend_schema(
    methods=["GET"],
    tags=["Staff/Memberships"],
    summary="멤버십 목록",
    responses=MembershipSerializer,
)
@extend_schema(
    methods=["POST"],
    tags=["Staff/Memberships"],
    summary="멤버십 생성",
    request=MembershipSerializer,
    responses={
        201: MembershipSerializer,
        400: OpenApiResponse(description="이미 해당 고객의 멤버십 존재"),
        403: OpenApiResponse(description="권한 없음 (OWNER/MANAGER만)"),
    }
)
class MembershipsView(APIView):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Membership.objects.all().order_by("customer_id")
        return Response(MembershipSerializer(qs, many=True).data)

    def post(self, request):
        if not IsOwnerOrManager().has_permission(request, self):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        ser = MembershipSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        if Membership.objects.filter(customer=ser.validated_data["customer"]).exists():
            return Response({"detail": "이미 해당 고객의 멤버십이 존재합니다."}, status=status.HTTP_400_BAD_REQUEST)
        obj = ser.save()
        return Response(MembershipSerializer(obj).data, status=status.HTTP_201_CREATED)


@extend_schema(
    tags=["Staff/Memberships"],
    summary="멤버십 단건 조회/수정/삭제",
    parameters=[OpenApiParameter("customer_id", int, OpenApiParameter.PATH, description="고객 ID")],
    responses={
        200: MembershipSerializer,
        204: OpenApiResponse(description="삭제 성공"),
        400: OpenApiResponse(description="customer 변경 불가 등"),
        403: OpenApiResponse(description="권한 없음 (수정/삭제)"),
        404: OpenApiResponse(description="존재하지 않음"),
    }
)
class MembershipDetailView(APIView):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get_object(self, customer_id: int) -> Membership:
        return get_object_or_404(Membership, customer_id=customer_id)

    def get(self, request, customer_id: int):
        obj = self.get_object(customer_id)
        return Response(MembershipSerializer(obj).data)

    def patch(self, request, customer_id: int):
        if not IsOwnerOrManager().has_permission(request, self):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        obj = self.get_object(customer_id)
        ser = MembershipSerializer(instance=obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        if "customer" in ser.validated_data and ser.validated_data["customer"].pk != obj.customer_id:
            return Response({"detail": "customer는 변경할 수 없습니다."}, status=status.HTTP_400_BAD_REQUEST)
        ser.save()
        return Response(MembershipSerializer(obj).data)

    def delete(self, request, customer_id: int):
        if not IsOwnerOrManager().has_permission(request, self):
            return Response({"detail": "권한이 없습니다."}, status=status.HTTP_403_FORBIDDEN)
        obj = self.get_object(customer_id)
        if hasattr(obj, "active"):
            obj.active = False
            obj.save(update_fields=["active"])
            return Response(status=status.HTTP_204_NO_CONTENT)
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ---------- Orders: 단건 상세 ----------
@extend_schema(
    tags=["Staff/Orders"],
    summary="주문 단건 상세(스태프용)",
    parameters=[OpenApiParameter("order_id", int, OpenApiParameter.PATH, description="주문 ID")],
    responses=StaffOrderDetailSerializer,
    examples=[OpenApiExample(
        name="응답 예시(요약)",
        value={
            "id": 123, "customer_id": 6, "ordered_at": "2025-10-28T10:10:10+09:00",
            "status": "pending", "order_source": "GUI",
            "receiver_name": "홍길동", "receiver_phone": "010-1111-2222",
            "delivery_address": "서울 중구 을지로 00",
            "geo_lat": "37.566000", "geo_lng": "126.978000", "place_label": "집",
            "address_meta": {"note": "경비실"},
            "card_last4": "4242",
            "subtotal_cents": 21000, "discount_cents": 1000, "total_cents": 20000,
            "meta": {"discounts":[{"type":"coupon","label":"WELCOME10","code":"WELCOME10","amount_cents":1000}]},
            "dinners": [],
            "coupons": [{"coupon": "WELCOME10","amount_cents": 1000,"channel": "GUI","redeemed_at": "2025-10-28T10:10:30+09:00"}],
            "membership": {"customer_id": 6,"percent_off": 5,"active": True,"valid_from": "2025-01-01T00:00:00+09:00","valid_until": None}
        },
        response_only=True
    )]
)
class StaffOrderDetailView(APIView):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request, order_id: int):
        order = (
            Order.objects
            .select_related("customer")
            .prefetch_related(
                "dinners__items__item",
                "dinners__items__options",
                "dinners__options",
                "coupon_redemptions__coupon",
            )
            .get(pk=order_id)
        )
        return Response(StaffOrderDetailSerializer(order).data, status=status.HTTP_200_OK)


# ---------- SSE (Orders) ----------
def _sse_headers(resp: StreamingHttpResponse) -> StreamingHttpResponse:
    resp["Content-Type"] = "text/event-stream; charset=utf-8"
    resp["Cache-Control"] = "no-cache, no-transform"
    return resp

@method_decorator(csrf_exempt, name="dispatch")
class OrdersSSEView(View):
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def dispatch(self, request, *args, **kwargs):
        # 수동 인증(StreamingHttpResponse + DRF 혼용 시)
        authed = False
        for cls in self.authentication_classes:
            inst = cls()
            res = inst.authenticate(request)
            if res:
                request.user, request.auth = res
                authed = True
                break
        if not authed:
            return HttpResponse("Unauthorized", status=401)
        for perm in self.permission_classes:
            if not perm().has_permission(request, self):
                return HttpResponse("Forbidden", status=403)
        return super().dispatch(request, *args, **kwargs)

    def _bootstrap(self, request):
        status_param = (request.GET.get("status") or "").strip()
        since_param = (request.GET.get("since") or "").strip()
        limit = int(request.GET.get("limit") or 20)
        limit = max(1, min(limit, 100))

        qs = Order.objects.all().order_by("-ordered_at")
        if status_param:
            statuses = [s.strip() for s in status_param.split(",") if s.strip()]
            qs = qs.filter(status__in=statuses)
        if since_param:
            dt = parse_datetime(since_param)
            if dt:
                qs = qs.filter(ordered_at__gte=dt)

        out = []
        for o in qs[:limit]:
            out.append({
                "id": o.id,
                "status": o.status,
                "ordered_at": o.ordered_at.isoformat(),
                "customer_id": o.customer_id,
                "order_source": o.order_source,
                "subtotal_cents": o.subtotal_cents,
                "total_cents": o.total_cents,
                "receiver_name": o.receiver_name,
                "place_label": o.place_label,
            })
        return out

    @extend_schema(
        tags=["Staff/SSE"],
        summary="주문 SSE 스트림",
        description=(
            "서버-전송 이벤트(SSE)로 주문 변화를 실시간 스트리밍합니다.\n\n"
            "**쿼리 파라미터**\n"
            "- `status`: 쉼표 구분 상태 필터(예: `pending,preparing`)\n"
            "- `since`: ISO8601로 이후 주문만 부트스트랩(예: `2025-10-28T00:00:00+09:00`)\n"
            "- `limit`: 부트스트랩 상한(1~100, 기본 20)\n\n"
            "**이벤트 타입**\n"
            "- `bootstrap`: 초기 주문 목록 배열\n"
            "- 그 외: `event` 필드명에 따라 도메인 이벤트(`orders_events`) 전송"
        ),
        parameters=[
            OpenApiParameter("status", str, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("since", str, OpenApiParameter.QUERY, required=False),
            OpenApiParameter("limit", int, OpenApiParameter.QUERY, required=False),
        ],
        responses={200: OpenApiResponse(description="text/event-stream (SSE)")},
        examples=[
            OpenApiExample(
                name="SSE 예시",
                value=(
                    "event: bootstrap\n"
                    "data: [{\"id\":123,\"status\":\"pending\",\"ordered_at\":\"2025-10-28T10:10:10+09:00\","
                    "\"customer_id\":6,\"order_source\":\"GUI\",\"subtotal_cents\":21000,\"total_cents\":20000,"
                    "\"receiver_name\":\"홍길동\",\"place_label\":\"집\"}]\n\n"
                    "event: orders_events\n"
                    "data: {\"event\":\"accepted\",\"order_id\":123,\"at\":\"2025-10-28T10:12:00+09:00\"}\n\n"
                ),
                response_only=True
            )
        ]
    )
    def get(self, request):
        def stream():
            yield "event: bootstrap\n"
            yield "data: " + json.dumps(self._bootstrap(request), ensure_ascii=False) + "\n\n"
            for msg in iter_order_notifications():
                name = msg.get("event", "message")
                yield f"event: {name}\n"
                yield "data: " + json.dumps(msg, ensure_ascii=False) + "\n\n"
        return _sse_headers(StreamingHttpResponse(stream()))


# ---------- Inventory (사람 주도) ----------
@extend_schema(
    methods=["GET"],
    tags=["Staff/Inventory"],
    summary="재고 아이템 목록/검색",
    parameters=[
        OpenApiParameter("q", str, OpenApiParameter.QUERY, required=False, description="이름/코드 부분검색"),
        OpenApiParameter("active", str, OpenApiParameter.QUERY, required=False, description="true/false"),
    ],
    responses=inline_serializer(
        name="InventoryListResp",
        fields={
            "count": serializers.IntegerField(),
            "items": inline_serializer(
                name="InventoryListItem",
                fields={
                    "code": serializers.CharField(),
                    "name": serializers.CharField(),
                    "active": serializers.BooleanField(),
                    "qty": serializers.IntegerField(),
                    "category": serializers.CharField(allow_null=True),
                    "soldout_reason": serializers.CharField(allow_null=True),
                    "price_cents": serializers.IntegerField(allow_null=True),
                    "updated_at": serializers.CharField(allow_null=True),
                }
            , many=True),
        }
    ),
    examples=[OpenApiExample(
        name="응답 예시",
        value={"count": 2, "items": [
            {"code": "KIMCHI", "name": "김치", "active": True, "qty": 12, "category": "Add-ons",
             "soldout_reason": None, "price_cents": 2000, "updated_at": "2025-10-28T10:00:00+09:00"},
            {"code": "COKE", "name": "코카콜라", "active": False, "qty": 0, "category": "Drinks",
             "soldout_reason": "소진", "price_cents": 1500, "updated_at": None}
        ]},
        response_only=True
    )]
)
@extend_schema(
    methods=["POST"],
    tags=["Staff/Inventory"],
    summary="재고 일괄 수정(사람 주도)",
    description=(
        "배치로 `qty`(절대값) 또는 `delta`(증감), `active`, `reason`를 반영합니다. "
        "`qty`와 `delta`를 동시에 전달하면 두 값을 순차적으로 적용합니다."
    ),
    request=inline_serializer(
        name="InventoryBulkUpdateReq",
        fields={
            "items": InventoryItemUpdateSerializer(many=True)
        }
    ),
    responses=inline_serializer(
        name="InventoryBulkUpdateResp",
        fields={"updated": inline_serializer(
            name="InventoryUpdatedItem",
            fields={
                "code": serializers.CharField(),
                "active": serializers.BooleanField(),
                "qty": serializers.IntegerField(),
            }, many=True
        )}
    ),
    examples=[
        OpenApiExample(
            name="요청 예시",
            value={"items": [
                {"code": "KIMCHI", "qty": 10},
                {"code": "COKE", "delta": -3, "active": False, "reason": "소진"}
            ]},
            request_only=True
        ),
        OpenApiExample(
            name="응답 예시",
            value={"updated": [{"code": "KIMCHI", "active": True, "qty": 10},
                               {"code": "COKE", "active": False, "qty": 0}]},
            response_only=True
        )
    ]
)
class InventoryItemsView(APIView):
    """
    GET  /api/staff/inventory/items      : 목록/검색
    POST /api/staff/inventory/items      : 일괄 수정 (qty/ delta / active / reason)
    자동 토글/방송 없음 — 사람 판단 우선
    """
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = MenuItem.objects.all().select_related("category")
        q = (request.query_params.get("q") or "").strip()
        if q:
            qs = qs.filter(name__icontains=q) | qs.filter(code__icontains=q)
        active_param = request.query_params.get("active")
        if active_param is not None:
            val = str(active_param).lower() in ("1","true","t","yes","y")
            qs = qs.filter(active=val)
        qs = qs.order_by("category__rank", "name")
        out = []
        for it in qs[:500]:
            attrs = dict(getattr(it, "attrs", {}) or {})
            qty = int(attrs.get("stock_qty") or 0)
            out.append({
                "code": it.code,
                "name": it.name,
                "active": it.active,
                "qty": qty,
                "category": getattr(getattr(it, "category", None), "name", None),
                "soldout_reason": attrs.get("soldout_reason"),
                "price_cents": getattr(it, "base_price_cents", None),
                "updated_at": getattr(it, "updated_at", None).isoformat() if getattr(it, "updated_at", None) else None,
            })
        return Response({"count": len(out), "items": out})

    def post(self, request):
        data = request.data or {}
        items = data.get("items")
        if not isinstance(items, list) or not items:
            return Response({"detail": "items 배열이 필요합니다."}, status=400)

        validated = []
        for idx, row in enumerate(items, start=1):
            s = InventoryItemUpdateSerializer(data=row)
            if not s.is_valid():
                return Response({"detail": f"items[{idx}] 유효하지 않습니다.", "errors": s.errors}, status=400)
            validated.append(s.validated_data)

        changed = []
        for row in validated:
            code = row["code"].strip()
            it = MenuItem.objects.filter(code=code).first()
            if not it:
                continue

            attrs = dict(getattr(it, "attrs", {}) or {})
            qty = int(attrs.get("stock_qty") or 0)

            if "qty" in row and row.get("qty") is not None:
                qty = max(0, int(row["qty"]))
            if "delta" in row and row.get("delta") is not None:
                qty = max(0, qty + int(row["delta"]))

            attrs["stock_qty"] = int(qty)

            if "active" in row:
                it.active = bool(row["active"])

            reason = row.get("reason")
            if reason is not None:
                attrs["soldout_reason"] = reason if not it.active or qty == 0 else None

            it.attrs = attrs
            if hasattr(it, "updated_at"):
                it.save(update_fields=["active", "attrs", "updated_at"])
            else:
                it.save(update_fields=["active", "attrs"])
            changed.append({"code": it.code, "active": it.active, "qty": attrs["stock_qty"]})
        return Response({"updated": changed}, status=200)


@extend_schema(
    methods=["PATCH"],
    tags=["Staff/Inventory"],
    summary="재고 단건 수정",
    parameters=[OpenApiParameter("code", str, OpenApiParameter.PATH, description="아이템 코드")],
    request=InventoryItemPartialUpdateSerializer,
    responses=inline_serializer(
        name="InventoryItemUpdateResp",
        fields={
            "code": serializers.CharField(),
            "active": serializers.BooleanField(),
            "qty": serializers.IntegerField(),
            "attrs": serializers.DictField(),
        }
    ),
    examples=[
        OpenApiExample(
            name="요청 예시(qty 절대값)",
            value={"qty": 7},
            request_only=True
        ),
        OpenApiExample(
            name="요청 예시(delta 증감)",
            value={"delta": -2, "reason": "소진"},
            request_only=True
        ),
        OpenApiExample(
            name="응답 예시",
            value={"code": "KIMCHI", "active": True, "qty": 7, "attrs": {"stock_qty": 7, "soldout_reason": None}},
            response_only=True
        ),
    ]
)
class InventoryItemDetailView(APIView):
    """단건 재고 수정: PATCH /api/staff/inventory/items/{code}"""
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]

    def patch(self, request, code: str):
        it = get_object_or_404(MenuItem, code=code)
        s = InventoryItemPartialUpdateSerializer(data=request.data, partial=True)
        s.is_valid(raise_exception=True)
        data = s.validated_data

        attrs = dict(getattr(it, "attrs", {}) or {})
        qty = int(attrs.get("stock_qty") or 0)

        if "qty" in data:
            qty = max(0, int(data["qty"]))
        if "delta" in data:
            qty = max(0, qty + int(data["delta"]))

        attrs["stock_qty"] = int(qty)

        if "active" in data:
            it.active = bool(data["active"])

        if "reason" in data:
            reason = data.get("reason")
            attrs["soldout_reason"] = reason if (reason is not None and (not it.active or qty == 0)) else attrs.get("soldout_reason")

        it.attrs = attrs
        if hasattr(it, "updated_at"):
            it.save(update_fields=["active", "attrs", "updated_at"])
        else:
            it.save(update_fields=["active", "attrs"])

        return Response({"code": it.code, "active": it.active, "qty": attrs["stock_qty"], "attrs": attrs})


@extend_schema(
    methods=["POST"],
    tags=["Staff/Inventory"],
    summary="재고 업로드(XLSX, 마감 후 조사 반영)",
    description=(
        "업로드된 **XLSX**의 각 행을 읽어 `code(또는 item_code)`, `qty(또는 quantity)`를 기준으로 "
        "수량을 **절대값으로** 반영합니다. 선택적으로 `active`, `reason`을 함께 반영할 수 있습니다.\n\n"
        "_참고: 현재 구현은 XLSX만 지원합니다._"
    ),
    request=inline_serializer(
        name="InventoryUploadReq",
        fields={
            "file": serializers.FileField(help_text="XLSX 파일")
        }
    ),
    responses={
        200: inline_serializer(
            name="InventoryUploadRespOK",
            fields={
                "updated": inline_serializer(
                    name="InventoryUploadUpdated",
                    fields={
                        "code": serializers.CharField(),
                        "qty": serializers.IntegerField(),
                        "active": serializers.BooleanField(),
                    }, many=True
                ),
                "errors": serializers.ListField(child=serializers.DictField(), default=[]),
            }
        ),
        400: inline_serializer(
            name="InventoryUploadRespError",
            fields={
                "updated": serializers.IntegerField(),
                "errors": serializers.ListField(child=serializers.DictField()),
            }
        ),
        207: inline_serializer(
            name="InventoryUploadRespMulti",
            fields={
                "updated": inline_serializer(
                    name="InventoryUploadUpdatedPartial",
                    fields={
                        "code": serializers.CharField(),
                        "qty": serializers.IntegerField(),
                        "active": serializers.BooleanField(),
                    }, many=True
                ),
                "errors": serializers.ListField(child=serializers.DictField()),
            }
        ),
    },
    examples=[
        OpenApiExample(
            name="성공 응답 예시",
            value={"updated": [{"code": "KIMCHI", "qty": 20, "active": True}], "errors": []},
            response_only=True
        ),
        OpenApiExample(
            name="에러 응답 예시(유효성 오류)",
            value={"updated": 0, "errors": [{"row": 2, "detail": "qty/quantity 누락"}]},
            response_only=True
        ),
    ]
)
class InventoryUploadView(APIView):
    """
    마감 후 재고조사 XLSX/CSV 업로드 → 수량 절대 반영
    필수 헤더: code|item_code, qty|quantity
    선택 헤더: active, reason
    """
    authentication_classes = [StaffJWTAuthentication]
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        f = request.FILES.get("file")
        if not f:
            return Response({"detail": "파일을 업로드 해주세요."}, status=400)

        filename = getattr(f, "name", "upload.bin").lower()
        rows, errors = [], []

        def norm(h): return str(h or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "")

        def push_row(d, idx):
            code = (d.get("code") or d.get("item_code") or "").strip()
            qty = d.get("qty") if d.get("qty") is not None else d.get("quantity")
            reason = d.get("reason")
            active = d.get("active")
            if not code:
                errors.append({"row": idx, "detail": "code 비어있음"}); return
            if qty is None:
                errors.append({"row": idx, "detail": "qty/quantity 누락"}); return
            try:
                qty = max(0, int(qty))
            except Exception:
                errors.append({"row": idx, "detail": f"qty 정수 아님: {qty}"}); return
            rows.append({"code": code, "qty": qty, "reason": reason, "active": active})

        # 안전하진 않음
        if filename.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception:
                return Response({"detail": "server error"}, status=500)
            wb = load_workbook(filename=io.BytesIO(f.read()), read_only=True, data_only=True)
            ws = wb.active
            headers = [norm(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            idx_map = {h: i for i, h in enumerate(headers)}
            for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                data = {}
                for key in ["code", "item_code", "qty", "quantity", "reason", "active"]:
                    pos = idx_map.get(norm(key))
                    if pos is not None and pos < len(row):
                        data[key] = row[pos].value
                push_row(data, r_idx)

        if errors:
            return Response({"updated": 0, "errors": errors}, status=400)

        updated = []
        for r in rows:
            it = MenuItem.objects.filter(code=r["code"]).first()
            if not it:
                errors.append({"code": r["code"], "detail": "해당 code의 MenuItem 없음"}); continue
            attrs = dict(getattr(it, "attrs", {}) or {})
            attrs["stock_qty"] = int(r["qty"])
            if r.get("reason") is not None:
                attrs["soldout_reason"] = r["reason"]
            if r.get("active") is not None:
                val = str(r["active"]).strip().lower()
                it.active = True if val in ("1", "true", "t", "yes", "y") else False
            it.attrs = attrs
            if hasattr(it, "updated_at"):
                it.save(update_fields=["attrs", "active", "updated_at"])
            else:
                it.save(update_fields=["attrs", "active"])
            updated.append({"code": it.code, "qty": attrs["stock_qty"], "active": it.active})

        status_code = 200 if not errors else 207
        return Response({"updated": updated, "errors": errors}, status=status_code)
